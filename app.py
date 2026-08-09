import os, sys, html, re
from pathlib import Path
from datetime import datetime
from flask import (

    Flask, render_template, request, jsonify, redirect,
    url_for, flash, session as flask_session, send_file
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, UserMixin, login_user, login_required,
    logout_user, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from sqlalchemy import text, func
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# SECRET_KEY MUST be provided by the hosting environment.
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY")
if not app.config["SECRET_KEY"]:
    raise RuntimeError(
        "SECRET_KEY must be configured. "
        "Generate one with: python -c \"import secrets; print(secrets.token_urlsafe(48))\""
    )

DB_USER = os.environ.get("DB_USER")
DB_PASS = os.environ.get("DB_PASS")
DB_HOST = os.environ.get("DB_HOST")
DB_NAME = os.environ.get("DB_NAME")

if not all([DB_USER, DB_PASS, DB_HOST, DB_NAME]):
    raise RuntimeError("DB_USER, DB_PASS, DB_HOST and DB_NAME must be configured.")

app.config["SQLALCHEMY_DATABASE_URI"] = (
    f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:5432/{DB_NAME}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_size": 10,
    "max_overflow": 20,
    "pool_timeout": 30,
    "pool_recycle": 1800,
    "pool_pre_ping": True,
}

# ---- Uploads / limits ----
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / "uploads"
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
app.config["UPLOAD_FOLDER"] = str(UPLOAD_FOLDER)

# Max upload size: 100 MB (matches the message shown in the UI)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024

# ---- Session cookie hardening ----
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true"
)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'


# --- SECURITY HEADERS ---
@app.after_request
def add_security_headers(resp):
    resp.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net "
        "https://unpkg.com "
        "https://esm.sh; "
        "style-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net; "
        "img-src 'self' data: blob:; "
        "media-src 'self' blob:; "
        "connect-src 'self' "
        "https://esm.sh "
        "https://unpkg.com "
        "https://cdn.jsdelivr.net "
        "https://fastly.jsdelivr.net; "
        "worker-src 'self' blob:; "
        "font-src 'self' data: "
        "https://cdn.jsdelivr.net;"
    )
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "SAMEORIGIN"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return resp



@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(error):
    return jsonify({
        "status": "error",
        "message": "The uploaded file is larger than the 100 MB limit."
    }), 413


# ---------------- MODELS ----------------
class User(UserMixin, db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)


class InventoryScan(db.Model):
    __tablename__ = "inventory_scan"

    id = db.Column(db.Integer, primary_key=True)

    code = db.Column(db.String(100), nullable=False, index=True)
    imei = db.Column(db.String(100), index=True)
    mac_address = db.Column(db.String(100), index=True)

    device_type = db.Column(db.String(100), nullable=False)
    department = db.Column(db.String(100))
    status = db.Column(db.String(50), nullable=False, index=True)
    substatus = db.Column(db.String(50))

    is_flagged = db.Column(db.Boolean, default=False, nullable=False, index=True)

    person_name = db.Column(db.String(100))
    employee_id = db.Column(db.String(50))
    email = db.Column(db.String(120))

    return_date = db.Column(db.Date)
    purchase_date = db.Column(db.Date)
    end_of_cycle = db.Column(db.Date)

    notes = db.Column(db.Text)

    # Store a short filename or URL here, NOT a full base64 image.
    image_data = db.Column(db.String(500))

    reason = db.Column(db.Text)

    timestamp = db.Column(
        db.DateTime, default=datetime.utcnow, nullable=False, index=True
    )


@login_manager.user_loader
def load_user(uid):
    return User.query.get(int(uid))


# ---------------- SCANNING RULES ----------------
DEVICE_TYPES = ["Laptop", "Mobile", "Monitor", "Printer", "Docking Station","Headset","Other"]
DEPARTMENTS  = ["IT", "FINANCE", "PROCUREMENT", "EMPLOYEE SERVICES" ,"Other"]
DEPARTMENT_ALIASES = {
    "HR": "EMPLOYEE SERVICES",
    "HUMAN RESOURCES": "EMPLOYEE SERVICES",
    "IT": "IT",
    "INFORMATION TECHNOLOGY": "IT",
    "FIN": "FINANCE",
    "PROC": "PROCUREMENT",
    "PROCUREMENTS": "PROCUREMENT",
}
STATUSES     = ["In Stock", "Loaned", "In Use", "Repair", "Retired"]

SUBSTATUS_RULES = {
    "In Stock": ["New", "Active"],
    "Loaned": ["Service Unit"],
    "In Use": ["Active"],
    "Repair": ["Ongoing"],
    "Retired": ["Lost", "End of Life"]
}

STATUS_FIELD_RULES = {
    "In Stock": ["purchase", "end"],
    "Loaned":   ["email", "date"],
    "In Use":   [],
    "Repair":   [],
    "Retired":  [],
}

DEVICE_IDENTIFIER_RULES = {
    "Laptop":  ["mac"],
    "Mobile":  ["imei", "mac"],
    "Monitor": [],
    "Printer": ["mac"],
    "Other":   [],
}

# Allowed characters in a scanned code. Loosen if your serials use spaces/symbols.
CODE_PATTERN = re.compile(r"[A-Za-z0-9._:\-\/]+")


# ---------------- INIT ----------------
def init_db():
    with app.app_context():
        try:
            db.create_all()

            try:
                db.session.execute(text(
                    'ALTER TABLE "user" '
                    'ADD COLUMN IF NOT EXISTS is_admin BOOLEAN DEFAULT FALSE'
                ))
                db.session.execute(text(
                    'ALTER TABLE inventory_scan ADD COLUMN IF NOT EXISTS reason TEXT'
                ))
                db.session.execute(text(
                    'ALTER TABLE inventory_scan '
                    'ADD COLUMN IF NOT EXISTS is_flagged BOOLEAN DEFAULT FALSE'
                ))
                db.session.execute(text(
                    'ALTER TABLE inventory_scan '
                    'ADD COLUMN IF NOT EXISTS substatus VARCHAR(50)'
                ))
                db.session.execute(text("""
                    CREATE INDEX IF NOT EXISTS ix_inventory_scan_code_timestamp
                    ON inventory_scan (code, timestamp DESC)
                """))
                db.session.execute(text("""
                    CREATE INDEX IF NOT EXISTS ix_inventory_scan_code_status
                    ON inventory_scan (code, status)
                """))
                db.session.execute(text("""
                    CREATE INDEX IF NOT EXISTS ix_inventory_scan_timestamp
                    ON inventory_scan (timestamp DESC)
                """))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Database migration error: {e}", file=sys.stderr)

            if not User.query.filter_by(username='admin').first():
                admin_password = os.environ.get("ADMIN_INITIAL_PASSWORD")
                if not admin_password:
                    raise RuntimeError("ADMIN_INITIAL_PASSWORD must be configured.")
                db.session.add(User(
                    username='admin',
                    password=generate_password_hash(admin_password),
                    is_admin=True
                ))
                db.session.commit()

        except Exception as e:
            print(f"DB Init Error: {e}", file=sys.stderr)


init_db()


def sanitize(val, length=100):
    if not val:
        return None
    return html.escape(str(val).strip()[:length])


def sanitize_title(val, length=100):
    cleaned = sanitize(val, length)
    if not cleaned:
        return None
    return cleaned.capitalize()


def parse_dt(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date() if s else None
    except Exception:
        return None


def compute_is_flagged(code, status, exclude_id=None):
    """
    Shared flag logic: a code becomes flagged once it has accumulated
    3 or more 'Repair' status records (including the one being saved).
    Any other status clears the flag.
    """
    if status != "Repair":
        return False

    query = InventoryScan.query.filter_by(code=code, status="Repair")
    if exclude_id is not None:
        query = query.filter(InventoryScan.id != exclude_id)

    repair_count = query.count()
    return (repair_count + 1) >= 3

def get_repair_count(code):
    """Return the total number of historical Repair events for a code."""
    return InventoryScan.query.filter(
        InventoryScan.code == code,
        InventoryScan.status == "Repair"
    ).count()


def refresh_asset_flag(code):
    """
    Recalculate and save the flag for every historical row
    belonging to the same asset code.
    """

    repair_count = InventoryScan.query.filter(
        InventoryScan.code == code,
        InventoryScan.status == "Repair"
    ).count()

    flagged = repair_count >= 3

    InventoryScan.query.filter(
        InventoryScan.code == code
    ).update(
        {
            InventoryScan.is_flagged: flagged
        },
        synchronize_session=False
    )

    return flagged, repair_count


def refresh_asset_flag(code):
    """
    Recalculate and save the flag for every historical row
    belonging to the same asset code.
    """
    repair_count = get_repair_count(code)
    flagged = repair_count >= 3

    InventoryScan.query.filter(
        InventoryScan.code == code
    ).update(
        {
            InventoryScan.is_flagged: flagged
        },
        synchronize_session=False
    )

    return flagged, repair_count




# ---------------- AUTH / PAGES ----------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = sanitize(request.form.get('username'))
        p = request.form.get('password')
        user = User.query.filter_by(username=u).first()
        if user and check_password_hash(user.password, p):
            login_user(user)
            return redirect(url_for('index'))
        flash('Invalid username or password')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    scans = InventoryScan.query.order_by(
        InventoryScan.timestamp.desc()
    ).limit(50).all()
    repair_counts = dict(
        db.session.query(InventoryScan.code, func.count(InventoryScan.id))
        .filter(InventoryScan.status == 'Repair')
        .group_by(InventoryScan.code).all()
    )
    flagged_codes = {c for c, n in repair_counts.items() if n >= 3}
    return render_template(
        'index.html', scans=scans, user=current_user,
        device_types=DEVICE_TYPES, departments=DEPARTMENTS,
        statuses=STATUSES, flagged_codes=flagged_codes
    )
  
@app.route('/session/start-mode', methods=['POST'])
@login_required
def session_start_mode():
    data = request.get_json() or {}

    scan_mode = sanitize(data.get('scanMode'))

    if scan_mode not in ["Single", "Multiple"]:
        return jsonify({
            "ok": False,
            "error": "Invalid scan mode."
        }), 400

    flask_session['scan_cfg'] = {
        "scanMode": scan_mode,
        "identifiers": []
    }

    return jsonify({
        "ok": True,
        "config": {
            "scanMode": scan_mode
        }
    })

 


# ---------------- SESSION ----------------
@app.route('/session/start', methods=['POST'])
@login_required
def session_start():
    d = request.get_json() or {}

    user = sanitize(d.get('user'))
    emp = sanitize(d.get('empId'))

    if not user or not emp:
        return jsonify({
            "ok": False,
            "error": "Employee Name and ID are mandatory."
        }), 400

    device = sanitize(d.get('device'))

    if device == 'Other':
        device = sanitize_title(
            d.get('otherDevice')
        ) or 'Other'

    dept = sanitize(d.get('dept'))

    if dept == 'Other':
        other_dept = sanitize(d.get('otherDept'))

        if not other_dept:
            return jsonify({
                "ok": False,
                "error": "Please specify a department."
            }), 400

        other_dept_upper = other_dept.upper()

        canonical = DEPARTMENT_ALIASES.get(
            other_dept_upper
        )

        existing_upper = [
            department.upper()
            for department in DEPARTMENTS
            if department != 'Other'
        ]

        if canonical:
            return jsonify({
                "ok": False,
                "error": (
                    f"'{other_dept_upper}' matches an existing "
                    f"department ('{canonical}'). "
                    "Please select it instead."
                )
            }), 400

        if other_dept_upper in existing_upper:
            return jsonify({
                "ok": False,
                "error": (
                    f"'{other_dept_upper}' already exists in the "
                    "department list. Please select it instead."
                )
            }), 400

        dept = other_dept.strip()

    status = sanitize(d.get('status'))

    if status not in STATUSES:
        return jsonify({
            "ok": False,
            "error": "Invalid status."
        }), 400

    valid_substatuses = SUBSTATUS_RULES.get(status, [])

    if status == "Loaned":
        substatus = "Service Unit"

    elif status == "Repair":
        substatus = "Ongoing"

    elif status == "In Use":
        substatus = "Active"

    elif valid_substatuses:
        substatus = sanitize(d.get('substatus'))

        if substatus not in valid_substatuses:
            return jsonify({
                "ok": False,
                "error": (
                    f"Substatus is required for {status}. "
                    f"Choose one of: "
                    f"{', '.join(valid_substatuses)}."
                )
            }), 400

    else:
        substatus = None

    base = (
        device
        if device in DEVICE_IDENTIFIER_RULES
        else 'Other'
    )

    scan_mode = sanitize(
        d.get('scanMode')
    ) or "Single"

    if scan_mode not in ["Single", "Multiple"]:
        return jsonify({
            "ok": False,
            "error": "Invalid scan mode."
        }), 400

    flask_session['scan_cfg'] = {
        "user": user,
        "empId": emp,
        "device": device,
        "dept": dept,
        "status": status,
        "substatus": substatus,
        "scanMode": scan_mode,
        "email": sanitize(d.get('email'), 120),
        "date": sanitize(d.get('date')),
        "purchase": sanitize(d.get('purchase')),
        "end": sanitize(d.get('end')),
        "notes": sanitize(d.get('notes'), 1000),
        "image_data": None,
        "identifiers": DEVICE_IDENTIFIER_RULES.get(
            base,
            []
        )
    }

    return jsonify({
        "ok": True,
        "config": {
            "user": user,
            "empId": emp,
            "device": device,
            "dept": dept,
            "status": status,
            "substatus": substatus,
            "scanMode": scan_mode,
            "requiredFields": STATUS_FIELD_RULES.get(
                status,
                []
            ),
            "identifiers": flask_session['scan_cfg'][
                'identifiers'
            ]
        }
    })


# ---------------- SCAN CHECK ----------------
@app.route('/scan/check', methods=['POST'])
@login_required
def scan_check():
    cfg = flask_session.get('scan_cfg', {
    "scanMode": "Single",
    "identifiers": []
})

    data = request.get_json() or {}

    code = sanitize(data.get("code"), length=100)

    if not code:
        return jsonify({
            "ok": False,
            "accept": False,
            "reason": "Empty code."
        }), 200

    if not CODE_PATTERN.fullmatch(code):
        return jsonify({
            "ok": False,
            "accept": False,
            "reason": "Invalid barcode characters."
        }), 200

    # Retrieve all previous tracking events.
    # Nothing is modified here.
    history = InventoryScan.query.filter_by(
        code=code
    ).order_by(
        InventoryScan.timestamp.desc()
    ).all()

    # Unknown serial.
    if not history:
        return jsonify({
            "ok": True,
            "accept": True,
            "exists": False,
            "code": code,
            "latest": None,
            "history": [],
            "nextIdentifiers": cfg.get("identifiers", []),
            "confirmMessage": (
                f"Serial '{code}' was not found in the database. "
                "Do you want to add it?"
            ),
            "requireReason": False,
            "flag": None
        })

    latest = history[0]

    repair_count = InventoryScan.query.filter_by(
        code=code,
        status="Repair"
    ).count()

    def serialize_scan(scan):
        return {
            "id": scan.id,
            "code": scan.code,
            "device_type": scan.device_type or "",
            "department": scan.department or "",
            "status": scan.status or "",
            "substatus": scan.substatus or "",
            "person_name": scan.person_name or "",
            "employee_id": scan.employee_id or "",
            "email": scan.email or "",
            "imei": scan.imei or "",
            "mac_address": scan.mac_address or "",
            "purchase_date": (
                scan.purchase_date.isoformat()
                if scan.purchase_date else ""
            ),
            "return_date": (
                scan.return_date.isoformat()
                if scan.return_date else ""
            ),
            "end_of_cycle": (
                scan.end_of_cycle.isoformat()
                if scan.end_of_cycle else ""
            ),
            "reason": scan.reason or "",
            "notes": scan.notes or "",
            "is_flagged": bool(scan.is_flagged),
            "timestamp": (
                scan.timestamp.isoformat()
                if scan.timestamp else ""
            )
        }

    return jsonify({
        "ok": True,
        "accept": True,
        "exists": True,
        "code": code,

        # Latest record is displayed as read-only information.
        "latest": serialize_scan(latest),

        # All previous events are displayed as history.
        "history": [
            serialize_scan(scan)
            for scan in history
        ],

        "nextIdentifiers": cfg.get("identifiers", []),
        "confirmMessage": None,
        "requireReason": False,
        "repair_count": repair_count,
        "is_flagged": repair_count >= 3,
        "flag": "red" if repair_count >= 3 else None
    })

# ---------------- SAVE TRACKING EVENT ----------------
@app.route('/scanned', methods=['POST'])
@login_required
def scanned():
    cfg = flask_session.get('scan_cfg')

    if not cfg:
        return jsonify({
            "status": "error",
            "message": "No active session."
        }), 400

    data = request.get_json() or {}

    code = sanitize(data.get("code"), length=100)

    if not code:
        return jsonify({
            "status": "error",
            "message": "Serial is required."
        }), 400

    if not CODE_PATTERN.fullmatch(code):
        return jsonify({
            "status": "error",
            "message": "Invalid barcode characters."
        }), 400

    # This value must be true when the code did not previously exist
    # and the user approved creating it.
    allow_new_record = data.get("allowNewRecord") is True

    latest = InventoryScan.query.filter_by(
        code=code
    ).order_by(
        InventoryScan.timestamp.desc()
    ).first()

    # If the code is new, require explicit confirmation.
    if latest is None and not allow_new_record:
        return jsonify({
            "status": "needs_confirmation",
            "message": (
                f"Serial '{code}' does not exist. "
                "Please confirm that you want to add it."
            )
        }), 409

    reason = sanitize(data.get("reason"), 500)

    # A reason is required when creating an In Use event
    # after the latest event was Retired.
    if (
        latest
        and latest.status == "Retired"
        and cfg["status"] == "In Use"
        and not reason
    ):
        return jsonify({
            "status": "error",
            "message": (
                "A reason is required to create an In Use "
                "tracking event for a Retired unit."
            )
        }), 400

    # Count previous repairs before adding this new event.
    repair_count = InventoryScan.query.filter_by(
        code=code,
        status="Repair"
    ).count()

    is_flagged = (
        cfg["status"] == "Repair"
        and (repair_count + 1) >= 3
    )

    try:
        # IMPORTANT:
        # Always INSERT a new row.
        # Never modify the previous row.
        tracking_event = InventoryScan(
            code=code,

            imei=sanitize(data.get("imei"), 100),
            mac_address=sanitize(data.get("mac_address"), 100),

            # These values come from the selected session.
            device_type=cfg["device"],
            department=cfg["dept"],
            status=cfg["status"],
            substatus=cfg["substatus"],

            person_name=cfg["user"],
            employee_id=cfg["empId"],
            email=cfg.get("email"),

            return_date=parse_dt(cfg.get("date")),
            purchase_date=parse_dt(cfg.get("purchase")),
            end_of_cycle=parse_dt(cfg.get("end")),

            notes=cfg.get("notes"),
            image_data=cfg.get("image_data"),

            reason=reason,
            is_flagged=is_flagged,

            # Explicit timestamp for the new event.
            timestamp=datetime.utcnow()
        )

        db.session.add(tracking_event)
        db.session.commit()

        return jsonify({
            "status": "success",
            "id": tracking_event.id,
            "is_flagged": tracking_event.is_flagged,
            "tracking_event": True
        })

    except Exception as error:
        db.session.rollback()

        print(
            f"Tracking event save error: {error}",
            file=sys.stderr
        )

        return jsonify({
            "status": "error",
            "message": "Unable to save tracking event."
        }), 500

# ---------------- EDIT: fetch ----------------
@app.route('/scan/<int:scan_id>', methods=['GET'])
@login_required
def get_scan(scan_id):
    s = InventoryScan.query.get(scan_id)
    if not s:
        return jsonify({"status": "error", "message": "Not found"}), 404

    return jsonify({
        "status": "success",
        "scan": {
            "id": s.id,
            "code": s.code,
            "imei": s.imei or "",
            "mac_address": s.mac_address or "",
            "device_type": s.device_type,
            "department": s.department or "",
            "status": s.status,
            "substatus": s.substatus or "",
            "person_name": s.person_name or "",
            "employee_id": s.employee_id or "",
            "email": s.email or "",
            "purchase_date": s.purchase_date.strftime('%Y-%m-%d') if s.purchase_date else "",
            "return_date": s.return_date.strftime('%Y-%m-%d') if s.return_date else "",
            "end_of_cycle": s.end_of_cycle.strftime('%Y-%m-%d') if s.end_of_cycle else "",
            "reason": s.reason or "",
            "notes": s.notes or "",
            "is_flagged": s.is_flagged
        }
    })
# ---------------- ADD: NEW HISTORICAL TRACKING EVENT ----------------
@app.route('/scan/<int:scan_id>/create-event', methods=['POST'])
@login_required
def create_tracking_event(scan_id):
    """
    Creates a new tracking event without modifying the previous row.
    All previous Repair events remain available for counting.
    """

    previous = InventoryScan.query.get(scan_id)

    if not previous:
        return jsonify({
            "status": "error",
            "message": "Record not found."
        }), 404

    data = request.get_json() or {}

    code = sanitize(data.get("code"), 100) or previous.code
    device_type = sanitize(data.get("device_type"), 100)
    department = sanitize(data.get("department"), 100)
    status_value = sanitize(data.get("status"), 50)

    if not code:
        return jsonify({
            "status": "error",
            "message": "Serial / Code is required."
        }), 400

    if not CODE_PATTERN.fullmatch(code):
        return jsonify({
            "status": "error",
            "message": "Invalid barcode characters."
        }), 400

    if not device_type:
        return jsonify({
            "status": "error",
            "message": "Device type is required."
        }), 400

    if status_value not in STATUSES:
        return jsonify({
            "status": "error",
            "message": "Invalid status."
        }), 400

    substatus_value = sanitize(data.get("substatus"), 50)
    valid_substatuses = SUBSTATUS_RULES.get(status_value, [])

    if status_value == "Loaned":
        substatus_value = "Service Unit"
    elif status_value == "Repair":
        substatus_value = "Ongoing"
    elif status_value == "In Use":
        substatus_value = "Active"
    elif valid_substatuses:
        if substatus_value not in valid_substatuses:
            return jsonify({
                "status": "error",
                "message": (
                    f"Substatus must be one of: "
                    f"{', '.join(valid_substatuses)}."
                )
            }), 400
    else:
        substatus_value = None

    reason = sanitize(data.get("reason"), 500)

    if (
        previous.status == "Retired"
        and status_value == "In Use"
        and not reason
    ):
        return jsonify({
            "status": "error",
            "message": (
                "A reason is required to reactivate "
                "a Retired unit."
            )
        }), 400

    try:
        # Count previous historical Repair events.
        previous_repair_count = get_repair_count(code)

        # Include this new event if it is Repair.
        repair_count = previous_repair_count
        
        if status_value == "Repair":
            repair_count += 1

        flagged = repair_count >= 3

        new_event = InventoryScan(
            code=code,
            imei=sanitize(data.get("imei"), 100),
            mac_address=sanitize(data.get("mac_address"), 100),

            device_type=device_type,
            department=department,
            status=status_value,
            substatus=substatus_value,

            person_name=sanitize(data.get("person_name"), 100),
            employee_id=sanitize(data.get("employee_id"), 50),
            email=sanitize(data.get("email"), 120),

            purchase_date=parse_dt(data.get("purchase_date")),
            return_date=parse_dt(data.get("return_date")),
            end_of_cycle=parse_dt(data.get("end_of_cycle")),

            reason=reason,
            notes=sanitize(data.get("notes"), 1000),

            is_flagged=(flagged or bool(previous.is_flagged)),
            timestamp=datetime.utcnow()
        )
            
        flagged, repair_count = refresh_asset_flag(code)
        new_event.is_flagged = flagged
        
        db.session.add(new_event)
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Tracking event created successfully.",
            "id": new_event.id,
            "repair_count": repair_count,
            "is_flagged": flagged
        })

    except Exception as error:
        db.session.rollback()

        print(
            f"Create tracking event error: {error}",
            file=sys.stderr
        )

        return jsonify({
            "status": "error",
            "message": "Unable to create tracking event."
        }), 500



# ---------------- SCAN: UPDATE EXISTING LATEST RECORD ----------------
@app.route('/scan/<int:scan_id>/update-existing', methods=['POST'])
@login_required
def update_existing_scan(scan_id):
    scan = InventoryScan.query.get(scan_id)

    if not scan:
        return jsonify({
            "status": "error",
            "message": "Record not found."
        }), 404

    data = request.get_json() or {}

    code = sanitize(data.get("code"), 100)

    if not code:
        return jsonify({
            "status": "error",
            "message": "Serial / Code is required."
        }), 400

    if not CODE_PATTERN.fullmatch(code):
        return jsonify({
            "status": "error",
            "message": "Invalid barcode characters."
        }), 400

    device_type = sanitize(data.get("device_type"), 100)
    department = sanitize(data.get("department"), 100)
    status_value = sanitize(data.get("status"), 50)

    if not device_type:
        return jsonify({
            "status": "error",
            "message": "Device type is required."
        }), 400

    if status_value not in STATUSES:
        return jsonify({
            "status": "error",
            "message": "Invalid status."
        }), 400

    valid_substatuses = SUBSTATUS_RULES.get(status_value, [])
    substatus_value = sanitize(data.get("substatus"), 50)

    if status_value == "Loaned":
        substatus_value = "Service Unit"

    elif status_value == "Repair":
        substatus_value = "Ongoing"

    elif status_value == "In Use":
        substatus_value = "Active"

    elif valid_substatuses:
        if substatus_value not in valid_substatuses:
            return jsonify({
                "status": "error",
                "message": (
                    f"Substatus must be one of: "
                    f"{', '.join(valid_substatuses)}."
                )
            }), 400

    else:
        substatus_value = None

    reason = sanitize(data.get("reason"), 500)

    # Preserve the existing reactivation validation rule.
    if scan.status == "Retired" and status_value == "In Use":
        if not reason:
            return jsonify({
                "status": "error",
                "message": (
                    "A reason is required to reactivate "
                    "a Retired unit."
                )
            }), 400

    try:
        # Update the existing latest row.
        # No new history row is inserted.
        scan.code = code
        scan.device_type = device_type
        scan.department = department
        scan.status = status_value
        scan.substatus = substatus_value

        scan.person_name = sanitize(
            data.get("person_name"), 100
        )
        scan.employee_id = sanitize(
            data.get("employee_id"), 50
        )
        scan.email = sanitize(
            data.get("email"), 120
        )

        scan.imei = sanitize(
            data.get("imei"), 100
        )
        scan.mac_address = sanitize(
            data.get("mac_address"), 100
        )

        scan.purchase_date = parse_dt(
            data.get("purchase_date")
        )
        scan.return_date = parse_dt(
            data.get("return_date")
        )
        scan.end_of_cycle = parse_dt(
            data.get("end_of_cycle")
        )

        scan.reason = reason
        scan.notes = sanitize(
            data.get("notes"), 1000
        )

        # Recalculate the flag while excluding this row.
        scan.is_flagged = compute_is_flagged(
            code,
            status_value,
            exclude_id=scan.id
        )

        scan.timestamp = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Record updated successfully.",
            "id": scan.id,
            "is_flagged": bool(scan.is_flagged)
        })

    except Exception as error:
        db.session.rollback()

        print(
            f"Existing record update error: {error}",
            file=sys.stderr
        )

        return jsonify({
            "status": "error",
            "message": "Unable to update the record."
        }), 500



# ---------------- ADMIN: ADD RECORD ----------------
@app.route('/scan/add', methods=['POST'])
@login_required
def add_scan():
    if not current_user.is_admin:
        return jsonify({
            "status": "error",
            "message": "Admin only."
        }), 403

    d = request.get_json() or {}

    code = sanitize(d.get("code"), 100)
    if not code:
        return jsonify({
            "status": "error",
            "message": "Serial / Code is required."
        }), 400

    if not CODE_PATTERN.fullmatch(code):
        return jsonify({
            "status": "error",
            "message": "Invalid barcode characters."
        }), 400

    device_type = sanitize(d.get("device_type"), 100)
    if not device_type or device_type == "Other":
        return jsonify({
            "status": "error",
            "message": "A valid device type is required."
        }), 400

    department = sanitize(d.get("department"), 100)

    # Disallow saving literal "Other" because the UI should send custom text.
    if department == "Other":
        return jsonify({
            "status": "error",
            "message": "Please specify a department instead of using Other."
        }), 400

    # Reject names that duplicate one of the canonical existing departments.
    if department:
        department_upper = department.upper()

        canonical_department = DEPARTMENT_ALIASES.get(department_upper)
        known_departments = {
            item.upper()
            for item in DEPARTMENTS
            if item != "Other"
        }

        if canonical_department:
            return jsonify({
                "status": "error",
                "message": (
                    f"'{department}' matches the existing department "
                    f"'{canonical_department}'. Select it from the list."
                )
            }), 400

        if department_upper in known_departments:
            return jsonify({
                "status": "error",
                "message": (
                    f"'{department}' already exists in the department list. "
                    "Select it from the list."
                )
            }), 400

        # Keep custom department format consistent with your scan session flow.
        if department_upper not in known_departments:
            department = department.strip()

    status_value = sanitize(d.get("status"), 50)
    if status_value not in STATUSES:
        return jsonify({
            "status": "error",
            "message": "Invalid status."
        }), 400

    substatus_value = sanitize(d.get("substatus"), 50)
    valid_substatuses = SUBSTATUS_RULES.get(status_value, [])

    if status_value == "Loaned":
        substatus_value = "Service Unit"

    elif status_value == "Repair":
        substatus_value = "Ongoing"

    elif status_value == "In Use":
        substatus_value = "Active"

    elif valid_substatuses and substatus_value not in valid_substatuses:
        return jsonify({
            "status": "error",
            "message": (
                f"Substatus must be one of: "
                f"{', '.join(valid_substatuses)}."
            )
        }), 400

    try:
        scan = InventoryScan(
            code=code,
            imei=sanitize(d.get("imei"), 100),
            mac_address=sanitize(d.get("mac_address"), 100),

            device_type=device_type,
            department=department,
            status=status_value,
            substatus=substatus_value,

            person_name=sanitize(d.get("person_name"), 100),
            employee_id=sanitize(d.get("employee_id"), 50),
            email=sanitize(d.get("email"), 120),

            purchase_date=parse_dt(d.get("purchase_date")),
            return_date=parse_dt(d.get("return_date")),
            end_of_cycle=parse_dt(d.get("end_of_cycle")),

            reason=sanitize(d.get("reason"), 500),
            notes=sanitize(d.get("notes"), 1000),

            is_flagged=compute_is_flagged(code, status_value)
        )

        db.session.add(scan)
        db.session.commit()

        return jsonify({
            "status": "success",
            "id": scan.id,
            "is_flagged": scan.is_flagged
        })

    except Exception as e:
        db.session.rollback()
        print(f"Admin add record error: {e}", file=sys.stderr)

        return jsonify({
            "status": "error",
            "message": "Unable to add the record."
        }), 500

# ---------------- EDIT: update ----------------
@app.route('/scan/<int:scan_id>/edit', methods=['POST'])
@login_required
def edit_scan(scan_id):
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Admin only"}), 403

    s = InventoryScan.query.get(scan_id)
    if not s:
        return jsonify({"status": "error", "message": "Not found"}), 404

    d = request.get_json() or {}

    status_value = sanitize(d.get('status'))
    if status_value not in STATUSES:
        return jsonify({"status": "error", "message": "Invalid status."}), 400

    valid_substatuses = SUBSTATUS_RULES.get(status_value, [])
    substatus_value = sanitize(d.get('substatus'))

    if status_value == "Loaned":
        substatus_value = "Service Unit"
    elif status_value == "Repair":
        substatus_value = "Ongoing"
    elif status_value == "In Use":
        substatus_value = "Active"
    elif valid_substatuses:
        if substatus_value not in valid_substatuses:
            return jsonify({
                "status": "error",
                "message": f"Substatus must be one of: {', '.join(valid_substatuses)}."
            }), 400
    else:
        substatus_value = None

    if status_value == "In Use" and s.status == "Retired":
        reason = sanitize(d.get('reason'), 500)
        if not reason:
            return jsonify({
                "status": "error",
                "message": "A reason is required to reactivate a Retired unit."
            }), 400
        s.reason = reason
    else:
        s.reason = sanitize(d.get('reason'), 500)

    new_code = sanitize(d.get('code')) or s.code

    # Recompute the flag against the (possibly new) code/status combo,
    # excluding this row itself from its own repair count.
    s.is_flagged = compute_is_flagged(new_code, status_value, exclude_id=s.id)

    try:
        s.code = new_code
        s.imei = sanitize(d.get('imei'))
        s.mac_address = sanitize(d.get('mac_address'))
        s.device_type = sanitize(d.get('device_type')) or s.device_type
        s.department = sanitize(d.get('department'), 100)
        s.status = status_value
        s.substatus = substatus_value
        s.person_name = sanitize(d.get('person_name'))
        s.employee_id = sanitize(d.get('employee_id'))
        s.email = sanitize(d.get('email'), 120)
        s.purchase_date = parse_dt(d.get('purchase_date'))
        s.return_date = parse_dt(d.get('return_date'))
        s.end_of_cycle = parse_dt(d.get('end_of_cycle'))
        s.notes = sanitize(d.get('notes'), 1000)

        db.session.commit()
        return jsonify({"status": "success", "id": s.id, "is_flagged": s.is_flagged})

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ---------------- ADMIN ----------------
@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
def manage_users():
    if not current_user.is_admin:
        return redirect(url_for('index'))
    if request.method == 'POST':
        u = sanitize(request.form.get('username'))
        p = request.form.get('password')
        a = bool(request.form.get('is_admin'))
        if User.query.filter_by(username=u).first():
            flash("User exists!")
        else:
            db.session.add(User(username=u, password=generate_password_hash(p), is_admin=a))
            db.session.commit()
            flash("User created.")
    return render_template('admin_users.html', users=User.query.all())


@app.route('/admin/users/delete/<int:user_id>')
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        return redirect(url_for('index'))
    u = User.query.get(user_id)
    if u and u.id != current_user.id:
        db.session.delete(u)
        db.session.commit()
    return redirect(url_for('manage_users'))


@app.route('/delete', methods=['POST'])
@login_required
def delete_scans():
    if not current_user.is_admin:
        return jsonify({"status": "error", "message": "Admin only"}), 403
    ids = (request.get_json() or {}).get('ids', [])
    if not ids:
        return jsonify({"status": "error", "message": "No ids provided."}), 400
    InventoryScan.query.filter(InventoryScan.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"status": "success"})


# ---------------- EXPORT ----------------
@app.route('/export/excel')
@login_required
def export_excel():
    scans = InventoryScan.query.order_by(InventoryScan.timestamp.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "Timestamp", "Serial/Code", "Device Type", "IMEI", "MAC Address",
        "Department", "Status", "Substatus", "Employee Name", "Employee ID",
        "Email", "Purchase Date", "Return Date", "End of Cycle", "Reason", "Notes"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for s in scans:
        ws.append([
            s.timestamp.strftime('%Y-%m-%d %H:%M:%S') if s.timestamp else "",
            s.code, s.device_type, s.imei or "", s.mac_address or "",
            s.department or "", s.status, s.substatus or "",
            s.person_name or "", s.employee_id or "", s.email or "",
            s.purchase_date.strftime('%Y-%m-%d') if s.purchase_date else "",
            s.return_date.strftime('%Y-%m-%d') if s.return_date else "",
            s.end_of_cycle.strftime('%Y-%m-%d') if s.end_of_cycle else "",
            s.reason or "", s.notes or ""
        ])

    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"inventory_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------- IMPORT ----------------
@app.route('/import/excel', methods=['POST'])
@login_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No file selected."}), 400

    file = request.files['file']
    if not file or not file.filename:
        return jsonify({"status": "error", "message": "Empty filename."}), 400

    filename = secure_filename(file.filename)
    if not filename.lower().endswith('.xlsx'):
        return jsonify({"status": "error", "message": "Only .xlsx files are allowed."}), 400

    unique_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{filename}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
    file.save(filepath)

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        imported = 0
        skipped = 0
        errors = []

        # ---- One query for all existing Repair counts, instead of a
        # per-row query inside compute_is_flagged(). ----
        existing_repair_counts = dict(
            db.session.query(InventoryScan.code, func.count(InventoryScan.id))
            .filter(InventoryScan.status == "Repair")
            .group_by(InventoryScan.code)
            .all()
        )
        # Tracks Repair rows seen so far in this batch, so multiple Repair
        # rows for the same code in one file still stack against the total.
        batch_repair_counts = {}

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if row is None or len(row) < 16:
                    skipped += 1
                    errors.append(f"Row {row_number}: Not enough columns.")
                    continue

                timestamp = row[0]
                serial = str(row[1]).strip() if row[1] else ""
                if not serial:
                    skipped += 1
                    errors.append(f"Row {row_number}: Serial Number is empty.")
                    continue

                status_value = str(row[6]).strip() if row[6] else "In Stock"
                substatus_value = str(row[7]).strip() if row[7] else None

                if status_value not in STATUSES:
                    skipped += 1
                    errors.append(f"Row {row_number}: Invalid status '{status_value}'.")
                    continue

                valid_substatuses = SUBSTATUS_RULES.get(status_value, [])

                if status_value == "Loaned":
                    substatus_value = "Service Unit"
                elif status_value == "Repair":
                    substatus_value = "Ongoing"
                elif status_value == "In Use":
                    substatus_value = "Active"
                elif valid_substatuses:
                    if substatus_value not in valid_substatuses:
                        skipped += 1
                        errors.append(
                            f"Row {row_number}: Invalid or missing substatus "
                            f"'{substatus_value}' for status '{status_value}'."
                        )
                        continue
                else:
                    substatus_value = None

                # ---- In-memory flag computation (no per-row DB query) ----
                if status_value == "Repair":
                    batch_repair_counts[serial] = batch_repair_counts.get(serial, 0) + 1
                    total_repairs = existing_repair_counts.get(serial, 0) + batch_repair_counts[serial]
                    is_flagged = total_repairs >= 3
                else:
                    is_flagged = False

                scan = InventoryScan(
                    code=serial,
                    device_type=row[2] or "Other",
                    imei=row[3],
                    mac_address=row[4],
                    department=row[5],
                    status=status_value,
                    substatus=substatus_value,
                    person_name=row[8],
                    employee_id=row[9],
                    email=row[10],
                    purchase_date=parse_dt(str(row[11])[:10]) if row[11] else None,
                    return_date=parse_dt(str(row[12])[:10]) if row[12] else None,
                    end_of_cycle=parse_dt(str(row[13])[:10]) if row[13] else None,
                    reason=row[14],
                    notes=row[15],
                    timestamp=(timestamp if isinstance(timestamp, datetime) else datetime.utcnow()),
                    is_flagged=is_flagged
                )
                db.session.add(scan)
                imported += 1

            except Exception as e:
                skipped += 1
                errors.append(f"Row {row_number}: {str(e)}")

        db.session.commit()
        wb.close()

        return jsonify({
            "status": "success",
            "imported": imported,
            "skipped": skipped,
            "errors": errors
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8506)
